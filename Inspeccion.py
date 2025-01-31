import streamlit as st
import pandas as pd
import os
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from fpdf import FPDF
from PIL import Image
from io import BytesIO
# Ruta del archivo Excel
BD = './DATA.xlsx'

# Cargar actividades iniciales si existen
df_ACTIVIDADES = pd.read_excel(BD, sheet_name="DATA")

# Configuración del Proyecto
st.sidebar.header("Configuración del Proyecto")
PROJECT_NAME = st.sidebar.text_input("Nombre del Proyecto", "MiProyecto")

IMAGE_FOLDER = os.path.join(PROJECT_NAME, "imagenes")
EXCEL_FILE = os.path.join(PROJECT_NAME, "actividades.xlsx")

# Crear carpetas si no existen
os.makedirs(IMAGE_FOLDER, exist_ok=True)

def load_or_create_excel():
    if not os.path.exists(EXCEL_FILE):
        df = pd.DataFrame(columns=["Fecha", "Actividad", "Descripción", "Imagenes"])
        df.to_excel(EXCEL_FILE, index=False)
    return pd.read_excel(EXCEL_FILE)

# Función para generar un archivo ZIP con imágenes y Excel
def create_zip():
    zip_filename = os.path.join(PROJECT_NAME, f"{PROJECT_NAME}_data.zip")
    
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Agregar el archivo Excel al ZIP
        if os.path.exists(EXCEL_FILE):
            zipf.write(EXCEL_FILE, os.path.basename(EXCEL_FILE))
        
        # Agregar todas las imágenes al ZIP
        for root, _, files in os.walk(IMAGE_FOLDER):
            for file in files:
                file_path = os.path.join(root, file)
                zipf.write(file_path, os.path.relpath(file_path, PROJECT_NAME))

    return zip_filename

def generate_unique_filename(original_name):
    """Genera un nombre único para la imagen basada en la fecha y hora"""
    name, ext = os.path.splitext(original_name)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"{name}_{timestamp}{ext}"

def save_data(actividad, descripcion, image_files):
    df = load_or_create_excel()
    existing_activity = df[df["Actividad"] == actividad]

    image_paths = existing_activity["Imagenes"].values[0].split(", ") if not existing_activity.empty and isinstance(existing_activity["Imagenes"].values[0], str) else []

    if image_files:
        for image_file in image_files:
            if image_file is not None:
                unique_name = generate_unique_filename(image_file.name)
                image_path = os.path.relpath(os.path.join(IMAGE_FOLDER, unique_name))
                with open(image_path, "wb") as f:
                    f.write(image_file.getbuffer())
                image_paths.append(image_path)

    images_string = ", ".join(image_paths) if image_paths else ""

    if not existing_activity.empty:
        df.loc[df["Actividad"] == actividad, "Imagenes"] = images_string
    else:
        new_data = pd.DataFrame({
            "Fecha": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            "Actividad": [actividad],
            "Descripción": [descripcion],
            "Imagenes": [images_string]
        })
        df = pd.concat([df, new_data], ignore_index=True)

    df.to_excel(EXCEL_FILE, index=False)

st.title("Registro de Actividades")
actividad = st.text_input("Nombre de la actividad")
descripcion = st.text_area("Descripción")

# Subir imágenes
image_files = st.file_uploader("Subir imágenes", type=["png", "jpg", "jpeg"], accept_multiple_files=True, key="file_uploader")

# Capturar imágenes con cámara
use_camera = st.checkbox("Tomar fotos desde la cámara")
if use_camera:
    camera_image = st.camera_input("Captura de cámara")
    if camera_image:
        image_files = image_files + [camera_image] if image_files else [camera_image]

if st.button("Guardar"):
    if actividad and descripcion:
        save_data(actividad, descripcion, image_files)
        st.success("Actividad guardada correctamente!")
    else:
        st.warning("Por favor, completa todos los campos.")

# Mostrar actividades registradas
with st.expander("Ver actividades registradas"):
    df = load_or_create_excel()
    st.dataframe(df.drop(columns=["Imagenes"]), use_container_width=True)

    for index, row in df.iterrows():
        st.write(f"### {row['Actividad']}")
        st.write(f"**Descripción:** {row['Descripción']}")
        if isinstance(row["Imagenes"], str):
            images = row["Imagenes"].split(", ")
            for img_path in images:
                if os.path.exists(img_path):
                    st.image(img_path, width=200)
                else:
                    st.write(f"Imagen no encontrada: {img_path}")

# Generar Informe PDF
def generate_pdf():
    df = load_or_create_excel()
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font('Arial', 'B', 16)
    pdf.cell(200, 10, txt=f"Informe del Proyecto: {PROJECT_NAME}", ln=True, align='C')
    pdf.ln(10)

    pdf.set_font('Arial', '', 12)
    for _, row in df.iterrows():
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(200, 10, txt=f"Actividad: {row['Actividad']}", ln=True)

        pdf.set_font('Arial', '', 12)
        pdf.multi_cell(0, 10, txt=f"Descripción: {row['Descripción']}")
        pdf.ln(5)

        imagenes = row["Imagenes"].split(", ") if isinstance(row["Imagenes"], str) else []
        for image_path in imagenes:
            if os.path.exists(image_path):
                img = Image.open(image_path)
                img_width, img_height = img.size
                aspect_ratio = img_height / img_width
                new_width = 100
                new_height = new_width * aspect_ratio
                pdf.image(image_path, w=new_width, h=new_height)
                pdf.ln(new_height + 5)
            else:
                pdf.cell(200, 10, txt=f"Imagen no encontrada: {image_path}", ln=True)

        pdf.ln(5)

    pdf_output = os.path.join(PROJECT_NAME, f"Informe_{PROJECT_NAME}.pdf")
    pdf.output(pdf_output)
    return pdf_output

if st.button("Generar Informe PDF"):
    pdf_file = generate_pdf()
    st.success(f"Informe PDF generado: {pdf_file}")
    with open(pdf_file, "rb") as f:
        st.download_button("Descargar Informe PDF", f, file_name=os.path.basename(pdf_file))



# Botón para descargar ZIP con imágenes y Excel
if st.button("Descargar ZIP con Datos e Imágenes"):
    zip_path = create_zip()
    st.success(f"Archivo ZIP generado: {zip_path}")
    
    with open(zip_path, "rb") as f:
        st.download_button("Descargar ZIP", f, file_name=os.path.basename(zip_path))